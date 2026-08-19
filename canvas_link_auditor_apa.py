import os
import re
import sys
import time
from datetime import datetime
from urllib.parse import urlparse, urljoin

import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURATION
# ============================================================

# Your Canvas URL.
# Examples:
#   https://youruniversity.instructure.com
CANVAS_URL = "https://YOUR-SCHOOL.instructure.com"

# Canvas course ID. Found at the end of the URL for your course.
# Example:
#   https://youruniversity.instructure.com/courses/123456
#                                                  ^^^^^^
COURSE_ID = "123456"

# Set this as an environment variable in PowerShell: DON'T put your token in this Python file!
# Example Powershell command:
# $env:CANVAS_API_TOKEN="your_token_here"
API_TOKEN = os.environ.get("CANVAS_API_TOKEN")

#Change the output filename if you'd like.
OUTPUT_FILE = "canvas_external_links.xlsx"

#Options for link checks and APA metadata retrieval. 
CHECK_LINKS = True
EXTRACT_APA_METADATA = True

REQUEST_DELAY = 0.1
HTTP_TIMEOUT = 15
USER_AGENT = "Canvas External Link Audit/1.0"


# ============================================================
# VALIDATION
# ============================================================

if "YOUR-SCHOOL" in CANVAS_URL:
    print("ERROR: Please set CANVAS_URL in the script.")
    sys.exit(1)

if "YOUR_COURSE_ID" in COURSE_ID:
    print("ERROR: Please set COURSE_ID in the script.")
    sys.exit(1)

if not API_TOKEN:
    print("ERROR: Canvas API token not found.")
    print('Set it first with: $env:CANVAS_API_TOKEN="your_token_here"')
    sys.exit(1)


BASE_URL = CANVAS_URL.rstrip("/")
API_BASE = f"{BASE_URL}/api/v1"

HEADERS = {
    "Authorization": f"Bearer {API_TOKEN}",
    "User-Agent": USER_AGENT,
}

WEB_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 Chrome/151.0 Safari/537.36 "
        "Canvas-Course-Link-Audit/1.0"
    )
}


# ============================================================
# CANVAS API
# ============================================================

def canvas_get(endpoint, params=None):
    url = f"{API_BASE}{endpoint}"
    results = []

    while url:
        response = requests.get(
            url,
            headers=HEADERS,
            params=params,
            timeout=HTTP_TIMEOUT,
        )

        if response.status_code != 200:
            print(f"Canvas API error {response.status_code}: {response.url}")
            response.raise_for_status()

        data = response.json()

        if isinstance(data, list):
            results.extend(data)
        else:
            return data

        url = None
        params = None

        if "Link" in response.headers:
            for link in response.headers["Link"].split(","):
                if 'rel="next"' in link:
                    match = re.search(r"<([^>]+)>", link)
                    if match:
                        url = match.group(1)

    return results


# ============================================================
# URL HELPERS
# ============================================================

def normalize_url(url):
    if not url:
        return ""

    url = url.strip().split("#")[0]
    parsed = urlparse(url)

    if parsed.path == "/":
        return url

    return url.rstrip("/")


def is_http_url(url):
    return bool(url) and url.lower().startswith(("http://", "https://"))


def is_canvas_internal(url):
    try:
        canvas_host = urlparse(BASE_URL).netloc.lower()
        link_host = urlparse(url).netloc.lower()
        return link_host == canvas_host or link_host.endswith("." + canvas_host)
    except Exception:
        return False


def extract_links_from_html(html):
    links = []

    if not html:
        return links

    soup = BeautifulSoup(html, "html.parser")

    for tag in soup.find_all("a", href=True):
        href = tag.get("href", "").strip()
        if not href:
            continue

        absolute_url = urljoin(BASE_URL, href)

        if is_http_url(absolute_url):
            links.append({
                "url": absolute_url,
                "link_text": tag.get_text(" ", strip=True),
            })

    # Include iframe/source links because these are often used
    # for embedded videos, simulations, and external resources.
    for tag in soup.find_all(["iframe", "img", "source"], src=True):
        src = tag.get("src", "").strip()
        if not src:
            continue

        absolute_url = urljoin(BASE_URL, src)

        if is_http_url(absolute_url):
            links.append({
                "url": absolute_url,
                "link_text": "",
            })

    return links


# ============================================================
# MODULES
# ============================================================

def get_modules():
    print("Retrieving modules...")

    modules = canvas_get(
        f"/courses/{COURSE_ID}/modules",
        params={
            "include[]": "items",
            "per_page": 100,
        },
    )

    print(f"  Found {len(modules)} modules.")
    return modules


def build_module_lookup():
    modules = get_modules()
    module_lookup = {}

    for module in modules:
        module_id = module.get("id")
        module_name = module.get("name", "")

        for item in module.get("items", []):
            item_type = item.get("type", "")
            content_id = item.get("content_id")

            key = (
                item_type,
                str(content_id) if content_id is not None else "",
            )

            module_lookup.setdefault(key, []).append({
                "id": module_id,
                "name": module_name,
            })

            # Direct external URL module item
            if item_type == "ExternalUrl" and item.get("external_url"):
                add_link(
                    item["external_url"],
                    "Module Item",
                    item.get("title", ""),
                    item.get("id", ""),
                    module_name,
                    module_id,
                    item.get("title", ""),
                    "ExternalUrl module item",
                )

            # External tool module item
            elif item_type == "ExternalTool" and item.get("external_url"):
                add_link(
                    item["external_url"],
                    "Module Item",
                    item.get("title", ""),
                    item.get("id", ""),
                    module_name,
                    module_id,
                    item.get("title", ""),
                    "ExternalTool module item",
                )

    return module_lookup


# ============================================================
# LINK RECORDS
# ============================================================

records = []


def add_link(
    url,
    content_type,
    content_name,
    content_id="",
    module_name="",
    module_id="",
    link_text="",
    source="",
):
    if not url or not is_http_url(url):
        return

    url = normalize_url(url)

    if not url or is_canvas_internal(url):
        return

    parsed = urlparse(url)

    records.append({
        "Module": module_name,
        "Module ID": module_id,
        "Content Type": content_type,
        "Content": content_name,
        "Content ID": content_id,
        "Link Text": link_text,
        "URL": url,
        "Domain": parsed.netloc.lower(),
        "Source": source,
    })


def scan_html(
    html,
    content_type,
    content_name,
    content_id="",
    module_name="",
    module_id="",
    source="Rich Text",
):
    for link in extract_links_from_html(html):
        add_link(
            link["url"],
            content_type,
            content_name,
            content_id,
            module_name,
            module_id,
            link["link_text"],
            source,
        )


# ============================================================
# CANVAS CONTENT SCANNERS
# ============================================================

def scan_pages(module_lookup):
    print("Scanning Pages...")

    pages = canvas_get(
        f"/courses/{COURSE_ID}/pages",
        params={
            "include[]": "body",
            "per_page": 100,
        },
    )

    print(f"  Found {len(pages)} pages.")

    for page in pages:
        name = page.get("title", "")
        page_id = page.get("page_id", "")

        modules = module_lookup.get(("Page", str(page_id)), [])

        if modules:
            for module in modules:
                scan_html(
                    page.get("body", ""),
                    "Page",
                    name,
                    page_id,
                    module["name"],
                    module["id"],
                    "Page body",
                )
        else:
            scan_html(
                page.get("body", ""),
                "Page",
                name,
                page_id,
                source="Page body",
            )


def scan_assignments(module_lookup):
    print("Scanning Assignments...")

    assignments = canvas_get(
        f"/courses/{COURSE_ID}/assignments",
        params={"per_page": 100},
    )

    print(f"  Found {len(assignments)} assignments.")

    for assignment in assignments:
        name = assignment.get("name", "")
        assignment_id = assignment.get("id", "")

        modules = module_lookup.get(("Assignment", str(assignment_id)), [])

        if modules:
            for module in modules:
                scan_html(
                    assignment.get("description", ""),
                    "Assignment",
                    name,
                    assignment_id,
                    module["name"],
                    module["id"],
                    "Assignment description",
                )
        else:
            scan_html(
                assignment.get("description", ""),
                "Assignment",
                name,
                assignment_id,
                source="Assignment description",
            )

        attrs = assignment.get("external_tool_tag_attributes")
        if attrs and attrs.get("url"):
            add_link(
                attrs["url"],
                "Assignment",
                name,
                assignment_id,
                source="External Tool",
            )


def scan_discussions(module_lookup):
    print("Scanning Discussions...")

    discussions = canvas_get(
        f"/courses/{COURSE_ID}/discussion_topics",
        params={"per_page": 100},
    )

    print(f"  Found {len(discussions)} discussions.")

    for discussion in discussions:
        name = discussion.get("title", "")
        discussion_id = discussion.get("id", "")

        modules = module_lookup.get(("Discussion", str(discussion_id)), [])

        if modules:
            for module in modules:
                scan_html(
                    discussion.get("message", ""),
                    "Discussion",
                    name,
                    discussion_id,
                    module["name"],
                    module["id"],
                    "Discussion message",
                )
        else:
            scan_html(
                discussion.get("message", ""),
                "Discussion",
                name,
                discussion_id,
                source="Discussion message",
            )


def scan_quizzes(module_lookup):
    print("Scanning Quizzes...")

    quizzes = canvas_get(
        f"/courses/{COURSE_ID}/quizzes",
        params={"per_page": 100},
    )

    print(f"  Found {len(quizzes)} quizzes.")

    for quiz in quizzes:
        name = quiz.get("title", "")
        quiz_id = quiz.get("id", "")

        modules = module_lookup.get(("Quiz", str(quiz_id)), [])

        if modules:
            for module in modules:
                scan_html(
                    quiz.get("description", ""),
                    "Quiz",
                    name,
                    quiz_id,
                    module["name"],
                    module["id"],
                    "Quiz description",
                )
        else:
            scan_html(
                quiz.get("description", ""),
                "Quiz",
                name,
                quiz_id,
                source="Quiz description",
            )


# ============================================================
# URL STATUS
# ============================================================

def check_url(url):
    try:
        response = requests.head(
            url,
            allow_redirects=True,
            timeout=HTTP_TIMEOUT,
            headers=WEB_HEADERS,
        )

        if response.status_code in [403, 405, 406]:
            response = requests.get(
                url,
                allow_redirects=True,
                timeout=HTTP_TIMEOUT,
                headers=WEB_HEADERS,
                stream=True,
            )

        return {
            "Status Code": response.status_code,
            "Status": (
                "OK"
                if 200 <= response.status_code < 400
                else "ERROR"
            ),
            "Final URL": response.url,
            "Redirected": normalize_url(response.url) != normalize_url(url),
            "Error": "",
        }

    except requests.exceptions.Timeout:
        return {
            "Status Code": "",
            "Status": "TIMEOUT",
            "Final URL": "",
            "Redirected": False,
            "Error": "Request timed out",
        }

    except requests.exceptions.SSLError as e:
        return {
            "Status Code": "",
            "Status": "SSL ERROR",
            "Final URL": "",
            "Redirected": False,
            "Error": str(e),
        }

    except requests.exceptions.RequestException as e:
        return {
            "Status Code": "",
            "Status": "ERROR",
            "Final URL": "",
            "Redirected": False,
            "Error": str(e),
        }


def check_all_links(df):
    if not CHECK_LINKS:
        return df

    print("\nChecking external URLs...\n")

    results = []

    for index, url in enumerate(df["URL"].drop_duplicates(), start=1):
        print(f"[{index}] {url}")

        result = check_url(url)
        results.append({"URL": url, **result})

        time.sleep(REQUEST_DELAY)

    return df.merge(pd.DataFrame(results), on="URL", how="left")


# ============================================================
# APA METADATA EXTRACTION
# ============================================================

def clean_text(value):
    if not value:
        return ""

    value = BeautifulSoup(str(value), "html.parser").get_text(" ", strip=True)
    value = re.sub(r"\s+", " ", value).strip()

    return value


def first_meta(soup, *names):
    for name in names:
        tag = soup.find("meta", attrs={"name": name})
        if tag and tag.get("content"):
            return clean_text(tag["content"])

        tag = soup.find("meta", attrs={"property": name})
        if tag and tag.get("content"):
            return clean_text(tag["content"])

    return ""


def parse_date(value):
    if not value:
        return ""

    value = value.strip()

    # ISO dates and common datetime forms
    candidates = [
        value,
        value.replace("Z", "+00:00"),
    ]

    for candidate in candidates:
        try:
            dt = datetime.fromisoformat(candidate)
            return str(dt.year)
        except Exception:
            pass

    # Four-digit year anywhere in metadata
    match = re.search(r"\b(19|20)\d{2}\b", value)
    if match:
        return match.group(0)

    return ""


def domain_to_site_name(domain):
    domain = domain.lower().split(":")[0]
    domain = re.sub(r"^www\.", "", domain)

    known = {
        "youtube.com": "YouTube",
        "youtu.be": "YouTube",
        "github.com": "GitHub",
        "microsoft.com": "Microsoft",
        "learn.microsoft.com": "Microsoft Learn",
        "docs.python.org": "Python Documentation",
        "python.org": "Python",
        "wikipedia.org": "Wikipedia",
        "developer.mozilla.org": "MDN Web Docs",
        "w3.org": "W3C",
        "labex.io": "LabEx",
        "testout.com": "TestOut",
    }

    for key, value in known.items():
        if domain == key or domain.endswith("." + key):
            return value

    parts = domain.split(".")
    if len(parts) >= 2:
        return parts[-2].capitalize()

    return domain


def determine_resource_type(url, soup):
    domain = urlparse(url).netloc.lower()
    path = urlparse(url).path.lower()

    if "youtube.com" in domain or "youtu.be" in domain:
        return "Video"

    if path.endswith(".pdf") or "application/pdf" in first_meta(
        soup, "og:type", "content-type"
    ).lower():
        return "PDF"

    if "github.com" in domain:
        return "Repository"

    if "learn.microsoft.com" in domain:
        return "Webpage"

    return "Webpage"


def fetch_apa_metadata(url):
    result = {
        "APA Author": "",
        "APA Date": "",
        "APA Title": "",
        "APA Website": "",
        "Resource Type": "",
        "APA Citation": "",
        "Citation Confidence": "Low",
        "Metadata Error": "",
    }

    try:
        response = requests.get(
            url,
            headers=WEB_HEADERS,
            timeout=HTTP_TIMEOUT,
            allow_redirects=True,
        )

        if response.status_code >= 400:
            result["Metadata Error"] = (
                f"HTTP {response.status_code}"
            )
            result["APA Website"] = domain_to_site_name(
                urlparse(url).netloc
            )
            return result

        soup = BeautifulSoup(response.text, "html.parser")

        title = (
            first_meta(
                soup,
                "citation_title",
                "dc.title",
                "og:title",
            )
            or clean_text(soup.title.get_text())
            if soup.title
            else ""
        )

        author = first_meta(
            soup,
            "citation_author",
            "author",
            "dc.creator",
            "article:author",
        )

        date_raw = first_meta(
            soup,
            "citation_publication_date",
            "citation_date",
            "date",
            "dc.date",
            "article:published_time",
            "datePublished",
            "dateCreated",
            "dateModified",
        )

        website = first_meta(
            soup,
            "og:site_name",
            "application-name",
        )

        final_url = response.url
        domain = urlparse(final_url).netloc

        if not website:
            website = domain_to_site_name(domain)

        date = parse_date(date_raw)
        resource_type = determine_resource_type(final_url, soup)

        if not author:
            # Use the site/organization as group author when
            # no individual author can be identified.
            author = website

        result.update({
            "APA Author": author,
            "APA Date": date,
            "APA Title": title,
            "APA Website": website,
            "Resource Type": resource_type,
        })

        # ----------------------------------------------------
        # APA 7 formatting
        # ----------------------------------------------------

        author_apa = author.rstrip(".")
        title_apa = title.rstrip(".")
        website_apa = website.rstrip(".")

        # Italicized title cannot be represented in plain text
        # reliably across Excel, so the citation uses plain text
        # and the Excel sheet separately identifies the title.
        if resource_type == "Video":
            citation = (
                f"{author_apa}. "
                f"({date if date else 'n.d.'}). "
                f"{title_apa} [Video]. "
                f"{website_apa}. "
                f"{final_url}"
            )

        elif resource_type == "PDF":
            citation = (
                f"{author_apa}. "
                f"({date if date else 'n.d.'}). "
                f"{title_apa} [PDF]. "
                f"{website_apa}. "
                f"{final_url}"
            )

        elif resource_type == "Repository":
            citation = (
                f"{author_apa}. "
                f"({date if date else 'n.d.'}). "
                f"{title_apa} [Computer software]. "
                f"{website_apa}. "
                f"{final_url}"
            )

        else:
            citation = (
                f"{author_apa}. "
                f"({date if date else 'n.d.'}). "
                f"{title_apa}. "
                f"{website_apa}. "
                f"{final_url}"
            )

        result["APA Citation"] = citation

        # Confidence assessment
        if author and title and date:
            result["Citation Confidence"] = "High"
        elif author and title:
            result["Citation Confidence"] = "Medium"
        elif title:
            result["Citation Confidence"] = "Low"
        else:
            result["Citation Confidence"] = "Very Low"

        return result

    except requests.exceptions.RequestException as e:
        result["Metadata Error"] = str(e)
        result["APA Website"] = domain_to_site_name(
            urlparse(url).netloc
        )
        return result

    except Exception as e:
        result["Metadata Error"] = str(e)
        return result


def add_apa_metadata(df):
    if not EXTRACT_APA_METADATA:
        return df

    print("\nRetrieving APA metadata...\n")

    metadata = []

    unique_urls = df["URL"].drop_duplicates().tolist()

    for index, url in enumerate(unique_urls, start=1):
        print(f"[APA {index}/{len(unique_urls)}] {url}")

        data = fetch_apa_metadata(url)
        metadata.append({
            "URL": url,
            **data,
        })

        time.sleep(REQUEST_DELAY)

    metadata_df = pd.DataFrame(metadata)

    return df.merge(
        metadata_df,
        on="URL",
        how="left",
    )


# ============================================================
# EXCEL FORMATTING
# ============================================================

def format_excel(filename):
    workbook = load_workbook(filename)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                "solid",
                fgColor="4472C4",
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:
                try:
                    max_length = max(
                        max_length,
                        len(str(cell.value)),
                    )
                except Exception:
                    pass

            worksheet.column_dimensions[column_letter].width = min(
                max_length + 2,
                70,
            )

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        headers = {
            cell.value: cell.column
            for cell in worksheet[1]
        }

        if "URL" in headers:
            url_col = headers["URL"]

            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row, url_col)

                if cell.value:
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"

        # APA citation column gets extra width
        if "APA Citation" in headers:
            citation_col = headers["APA Citation"]
            worksheet.column_dimensions[
                get_column_letter(citation_col)
            ].width = 90

    workbook.save(filename)


# ============================================================
# REPORT
# ============================================================

def create_report():
    print("=" * 65)
    print("CANVAS EXTERNAL LINK + APA CITATION AUDIT")
    print("=" * 65)
    print()

    module_lookup = build_module_lookup()

    scan_pages(module_lookup)
    scan_assignments(module_lookup)
    scan_discussions(module_lookup)
    scan_quizzes(module_lookup)

    if not records:
        print("\nNo external links were found.")
        return

    df = pd.DataFrame(records)

    # Status checking
    df = check_all_links(df)

    # APA metadata
    df = add_apa_metadata(df)

    # Sort master data
    df = df.sort_values(
        by=["Module", "Content Type", "Content", "URL"],
        na_position="last",
    )

    # Unique URL report
    unique_df = (
        df.groupby("URL", as_index=False)
        .agg({
            "Domain": "first",
            "APA Author": "first",
            "APA Date": "first",
            "APA Title": "first",
            "APA Website": "first",
            "Resource Type": "first",
            "APA Citation": "first",
            "Citation Confidence": "first",
            "Status Code": "first",
            "Status": "first",
            "Final URL": "first",
            "Redirected": "first",
            "Error": "first",
            "Metadata Error": "first",
            "Module": lambda x: "; ".join(
                sorted(
                    set(
                        str(v)
                        for v in x
                        if str(v) != "nan"
                        and str(v).strip()
                    )
                )
            ),
            "Content Type": lambda x: "; ".join(
                sorted(set(x))
            ),
        })
    )

    # Domain summary
    domains_df = (
        df.groupby("Domain")
        .agg(
            Links=("URL", "count"),
            Unique_URLs=("URL", "nunique"),
        )
        .reset_index()
        .sort_values(
            "Unique_URLs",
            ascending=False,
        )
    )

    # Broken links
    broken_df = df[
        ~df["Status"].isin(["OK"])
    ].copy()

    # Citation review
    citation_review_df = df[
        df["Citation Confidence"].isin(
            ["Low", "Very Low"]
        )
    ].copy()

    # Summary
    summary_df = pd.DataFrame({
        "Metric": [
            "Canvas URL",
            "Course ID",
            "Report Generated",
            "Total external link occurrences",
            "Unique external URLs",
            "Unique external domains",
            "Working URLs",
            "Problem URLs",
            "High-confidence APA citations",
            "Medium-confidence APA citations",
            "Low-confidence APA citations",
            "Very-low-confidence APA citations",
        ],
        "Value": [
            BASE_URL,
            COURSE_ID,
            pd.Timestamp.now(),
            len(df),
            df["URL"].nunique(),
            df["Domain"].nunique(),
            len(df[df["Status"] == "OK"]),
            len(broken_df),
            len(df[df["Citation Confidence"] == "High"]),
            len(df[df["Citation Confidence"] == "Medium"]),
            len(df[df["Citation Confidence"] == "Low"]),
            len(df[df["Citation Confidence"] == "Very Low"]),
        ],
    })

    # APA citations sheet
    apa_df = unique_df[
        [
            "Module",
            "Content Type",
            "URL",
            "APA Author",
            "APA Date",
            "APA Title",
            "APA Website",
            "Resource Type",
            "APA Citation",
            "Citation Confidence",
        ]
    ].copy()

    print(f"\nCreating {OUTPUT_FILE}...")

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl",
    ) as writer:

        summary_df.to_excel(
            writer,
            sheet_name="Summary",
            index=False,
        )

        df.to_excel(
            writer,
            sheet_name="All External Links",
            index=False,
        )

        unique_df.to_excel(
            writer,
            sheet_name="Unique URLs",
            index=False,
        )

        apa_df.to_excel(
            writer,
            sheet_name="APA Citations",
            index=False,
        )

        domains_df.to_excel(
            writer,
            sheet_name="Domains",
            index=False,
        )

        broken_df.to_excel(
            writer,
            sheet_name="Broken Links",
            index=False,
        )

        citation_review_df.to_excel(
            writer,
            sheet_name="Citation Review",
            index=False,
        )

        df.sort_values(
            by=["Module", "Content", "URL"],
            na_position="last",
        ).to_excel(
            writer,
            sheet_name="By Module",
            index=False,
        )

    format_excel(OUTPUT_FILE)

    print()
    print("=" * 65)
    print("COMPLETE")
    print("=" * 65)
    print(f"Report: {OUTPUT_FILE}")
    print(f"Total links: {len(df)}")
    print(f"Unique URLs: {df['URL'].nunique()}")
    print(f"Domains: {df['Domain'].nunique()}")
    print(f"Working URLs: {len(df[df['Status'] == 'OK'])}")
    print(f"Problem URLs: {len(broken_df)}")
    print(f"High-confidence citations: {len(df[df['Citation Confidence'] == 'High'])}")
    print(f"Needs citation review: {len(citation_review_df)}")
    print()


if __name__ == "__main__":
    try:
        create_report()
    except KeyboardInterrupt:
        print("\nScan cancelled.")
    except Exception as e:
        print("\nERROR:")
        print(e)
        raise
